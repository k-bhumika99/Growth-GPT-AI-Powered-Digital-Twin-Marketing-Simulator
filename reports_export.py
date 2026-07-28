"""
reports_export.py — Reporting & Export Module
Builds downloadable CSV, Excel (.xlsx), and PDF reports for saved campaigns,
plus bulk exports across all campaigns for the Dashboard/Reports pages.
"""
import io
import csv
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, ListFlowable, ListItem
)

BRAND_PRIMARY = colors.HexColor("#D6725C")
BRAND_VIOLET = colors.HexColor("#8E7CC3")
BRAND_INK = colors.HexColor("#2B2418")
BRAND_MUTED = colors.HexColor("#7A6E58")
BRAND_CREAM = colors.HexColor("#FBF3E6")

def _twins(record):
    return record["result"].get("digital_twins", []) or []

def _sections(record):
    result = record["result"]
    ca = result.get("campaign_analysis", {}) or {}
    predictions = result.get("predictions", {}) or {}
    readiness = result.get("readiness", {}) or {}
    gp = result.get("growth_prediction", {}) or {}
    
    sb = predictions.get("sentiment_breakdown") or gp.get("sentiment_breakdown") or {}
    
    def get_val(node, key):
        if not node:
            return None
        val = node.get(key)
        if isinstance(val, dict):
            return val.get("value")
        return val

    engagement_score = get_val(predictions, "engagement_score") or gp.get("engagement_score") or 70
    conversion_probability = get_val(predictions, "conversion_probability") or gp.get("conversion_probability") or 50
    growth_potential = readiness.get("risk_level") or gp.get("growth_potential") or "Medium"
    growth_potential_reason = readiness.get("launch_recommendation") or gp.get("growth_potential_reason") or "Tweak campaign."

    resolved_gp = {
        "engagement_score": engagement_score,
        "conversion_probability": conversion_probability,
        "growth_potential": growth_potential,
        "growth_potential_reason": growth_potential_reason,
        "sentiment_breakdown": sb,
        "campaign_performance_score": get_val(predictions, "campaign_performance_score") or 65,
        "virality_potential": get_val(predictions, "virality_potential") or 35,
        "estimated_roi": get_val(predictions, "estimated_roi") or "2.0x",
        "customer_retention_potential": get_val(predictions, "customer_retention_potential") or 55,
    }

    improvements = result.get("improvements", {}) or {}
    cat_rec = improvements.get("categorized_recommendations", {}) or {}
    
    rec = {
        "priority_action": readiness.get("launch_recommendation") or (cat_rec.get("high")[0] if cat_rec.get("high") else "Optimize visual contrast"),
        "weak_points": cat_rec.get("high") or ["Weak headline presentation"],
        "headline_suggestions": [improvements.get("headline")] if improvements.get("headline") else ["Unlock New Potential Now"],
        "cta_suggestions": [improvements.get("cta")] if improvements.get("cta") else ["Claim Offer Now"],
        "targeting_strategy": improvements.get("platform_strategy") or "Leverage platform organic video content.",
    }

    return ca, resolved_gp, rec

# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def campaign_csv(record) -> bytes:
    ca, gp, rec = _sections(record)
    sb = gp.get("sentiment_breakdown", {}) or {}
    result = record["result"]
    psych = result.get("psychology_analysis", {}) or {}
    img_anal = result.get("image_analysis", {}) or {}

    buf = io.StringIO()
    w = csv.writer(buf)

    w.writerow(["Growth GPT — Enterprise Campaign Report"])
    w.writerow(["Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")])
    w.writerow([])
    w.writerow(["Campaign Overview"])
    w.writerow(["Campaign/Product Name", record.get("product_name")])
    w.writerow(["Objective", record.get("objective")])
    w.writerow(["Budget Context", record.get("budget") or "—"])
    w.writerow(["Created", record.get("created_at")])
    w.writerow(["Tone", ca.get("tone_analysis") or "—"])
    w.writerow(["Core Objective Summary", ca.get("summary") or "—"])
    w.writerow([])
    w.writerow(["Predictions & Performance Scorecard"])
    w.writerow(["Engagement Score", gp.get("engagement_score")])
    w.writerow(["Conversion Probability (%)", gp.get("conversion_probability")])
    w.writerow(["Growth Potential / Risk Level", gp.get("growth_potential")])
    w.writerow(["Launch Recommendation", gp.get("growth_potential_reason")])
    w.writerow(["Estimated ROI", gp.get("estimated_roi")])
    w.writerow(["Sentiment — Positive (%)", sb.get("positive")])
    w.writerow(["Sentiment — Neutral (%)", sb.get("neutral")])
    w.writerow(["Sentiment — Negative (%)", sb.get("negative")])
    w.writerow([])
    w.writerow(["Marketing Psychology Ratings"])
    for k, v in psych.items():
        if isinstance(v, (int, float)):
            w.writerow([k.title(), v])
    w.writerow([])
    w.writerow(["Image Analysis / Ad Evaluation"])
    w.writerow(["Visual Score", img_anal.get("visual_score") or "—"])
    w.writerow(["Text Readability", img_anal.get("text_readability") or "—"])
    w.writerow(["CTA Visibility", img_anal.get("cta_visibility") or "—"])
    w.writerow(["Brand Visibility", img_anal.get("brand_visibility") or "—"])
    w.writerow([])
    w.writerow(["Digital Twins & Behaviour Simulation"])
    w.writerow(["Name", "Age", "Occupation", "Income", "Buying Behaviour", "Buying Motivation", "Buying Trigger", "First Impression", "Objection", "Decision"])
    for t in _twins(record):
        sim = t.get("simulation", {}) or {}
        w.writerow([
            t.get("name"), t.get("age"), t.get("occupation"), t.get("income"),
            t.get("buying_behaviour"), t.get("buying_motivation"), t.get("buying_trigger"),
            sim.get("first_impression"), sim.get("questions", ["—"])[0], sim.get("buying_decision")
        ])
    return buf.getvalue().encode("utf-8-sig")


def all_campaigns_csv(rows) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Campaign", "Objective", "Engagement Score", "Conversion Probability (%)",
                "Growth Potential", "Created"])
    for c in rows:
        w.writerow([c.get("product_name"), c.get("objective"), c.get("engagement_score"),
                    c.get("conversion_probability"), c.get("growth_potential"), c.get("created_at")])
    return buf.getvalue().encode("utf-8-sig")

# ---------------------------------------------------------------------------
# Excel (.xlsx)
# ---------------------------------------------------------------------------

def _style_header(ws, row_idx, ncols, fill="D6725C"):
    fill_style = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill_style
        cell.alignment = Alignment(vertical="center", wrap_text=True)

def campaign_xlsx(record) -> bytes:
    ca, gp, rec = _sections(record)
    sb = gp.get("sentiment_breakdown", {}) or {}
    result = record["result"]
    psych = result.get("psychology_analysis", {}) or {}
    img_anal = result.get("image_analysis", {}) or {}
    wb = Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Growth GPT — Campaign Report"])
    ws["A1"].font = Font(bold=True, size=14, color="2B2418")
    ws.append(["Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")])
    ws.append([])
    summary_rows = [
        ("Campaign", record.get("product_name")),
        ("Objective", record.get("objective")),
        ("Budget", record.get("budget") or "—"),
        ("Created", record.get("created_at")),
        ("Core Summary", ca.get("summary")),
        ("Tone", ca.get("tone_analysis")),
        ("Engagement Score", gp.get("engagement_score")),
        ("Conversion Probability (%)", gp.get("conversion_probability")),
        ("Growth Potential", gp.get("growth_potential")),
        ("Launch Recommendation", gp.get("growth_potential_reason")),
        ("Estimated ROI", gp.get("estimated_roi")),
        ("Sentiment — Positive (%)", sb.get("positive")),
        ("Sentiment — Neutral (%)", sb.get("neutral")),
        ("Sentiment — Negative (%)", sb.get("negative")),
    ]
    for label, val in summary_rows:
        ws.append([label, val])
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 60

    # --- Digital twins sheet ---
    ws2 = wb.create_sheet("Digital Twins")
    headers = ["Name", "Age", "Occupation", "Income", "Buying Behaviour", "Platform",
               "Buying Motivation", "Buying Trigger", "First Impression", "Decision"]
    ws2.append(headers)
    _style_header(ws2, 1, len(headers))
    for t in _twins(record):
        sim = t.get("simulation", {}) or {}
        ws2.append([
            t.get("name"), t.get("age"), t.get("occupation"), t.get("income"),
            t.get("buying_behaviour"), t.get("preferred_platform"),
            t.get("buying_motivation"), t.get("buying_trigger"),
            sim.get("first_impression"), sim.get("buying_decision")
        ])
    widths = [20, 6, 20, 14, 30, 14, 25, 20, 30, 12]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # --- Psychology Sheet ---
    ws3 = wb.create_sheet("Psychology & Image")
    ws3.append(["Psychological Triggers", "Rating"])
    _style_header(ws3, 1, 2, fill="8E7CC3")
    for k, v in psych.items():
        if isinstance(v, (int, float)):
            ws3.append([k.title(), v])
    ws3.append([])
    ws3.append(["Ad Image Evaluation", "Value"])
    _style_header(ws3, ws3.max_row, 2, fill="8E7CC3")
    for k, v in img_anal.items():
        if isinstance(v, (str, int, float)) and k != "strengths" and k != "weaknesses" and k != "suggestions":
            ws3.append([k.title().replace("_", " "), v])
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def all_campaigns_xlsx(rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Campaigns"
    headers = ["Campaign", "Objective", "Engagement Score", "Conversion Probability (%)",
               "Growth Potential", "Created"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    for c in rows:
        ws.append([c.get("product_name"), c.get("objective"), c.get("engagement_score"),
                   c.get("conversion_probability"), c.get("growth_potential"), c.get("created_at")])
    widths = [26, 16, 16, 20, 16, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def campaign_pdf(record) -> bytes:
    ca, gp, rec = _sections(record)
    sb = gp.get("sentiment_breakdown", {}) or {}
    result = record["result"]
    psych = result.get("psychology_analysis", {}) or {}
    img_anal = result.get("image_analysis", {}) or {}

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=22 * mm, bottomMargin=18 * mm, leftMargin=18 * mm, rightMargin=18 * mm,
        title=f"Growth GPT Report — {record.get('product_name', '')}",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleBrand", parent=styles["Title"], textColor=BRAND_INK, fontSize=20)
    h2_style = ParagraphStyle("H2Brand", parent=styles["Heading2"], textColor=BRAND_PRIMARY, spaceBefore=14, spaceAfter=8, fontSize=13)

    body_style = ParagraphStyle("BodyBrand", parent=styles["BodyText"], textColor=BRAND_INK, fontSize=9.5, leading=13.5)
    muted_style = ParagraphStyle("MutedBrand", parent=styles["BodyText"], textColor=BRAND_MUTED, fontSize=9)

    story = []
    story.append(Paragraph("Growth GPT — Enterprise Simulation Report", title_style))
    story.append(Paragraph(f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", muted_style))
    story.append(Spacer(1, 10))

    meta_table = Table([
        ["Campaign Name", record.get("product_name") or "—"],
        ["Objective", record.get("objective") or "—"],
        ["Budget Context", record.get("budget") or "—"],
        ["Created", (record.get("created_at") or "")[:19].replace("T", " ")],
        ["Tone", ca.get("tone_analysis") or "—"],
        ["Core Summary", ca.get("summary") or "—"],
    ], colWidths=[120, 350])
    meta_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), BRAND_MUTED),
        ("TEXTCOLOR", (1, 0), (1, -1), BRAND_INK),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.HexColor("#E7DCC4")),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 8))

    story.append(Paragraph("Campaign Predictions", h2_style))
    metrics_table = Table([
        ["Engagement Score", f"{gp.get('engagement_score', '—')}/100"],
        ["Conversion Probability", f"{gp.get('conversion_probability', '—')}%"],
        ["Risk / Readiness Level", gp.get("growth_potential") or "—"],
        ["Estimated ROI", gp.get("estimated_roi") or "—"],
        ["Sentiment (Pos / Neu / Neg)", f"{sb.get('positive', '—')}% / {sb.get('neutral', '—')}% / {sb.get('negative', '—')}%"],
    ], colWidths=[180, 290])
    metrics_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, -1), BRAND_CREAM),
        ("TEXTCOLOR", (0, 0), (0, -1), BRAND_MUTED),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#E7DCC4")),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E7DCC4")),
    ]))
    story.append(metrics_table)
    story.append(Spacer(1, 8))

    story.append(Paragraph("Digital Twins Reactions Summary", h2_style))
    for t in _twins(record)[:5]: # include top 5 twins in PDF to keep pages reasonable
        sim = t.get("simulation", {}) or {}
        name_line = f"<b>{t.get('name', '—')}</b> ({t.get('age', '—')}) — {t.get('occupation', '—')}"
        story.append(Paragraph(name_line, body_style))
        story.append(Paragraph(
            f"Impression: {sim.get('first_impression', '—')} · Decision: <b>{sim.get('buying_decision', '—')}</b>", muted_style))
        if sim.get("comment"):
            story.append(Paragraph(f"“{sim['comment']}”", body_style))
        story.append(Spacer(1, 6))

    story.append(Paragraph("Image Optimization Insights", h2_style))
    story.append(Paragraph(f"<b>Visual Layout Score:</b> {img_anal.get('visual_score', '—')}/100", body_style))
    story.append(Paragraph(f"Readability: {img_anal.get('text_readability', '—')} · CTA: {img_anal.get('cta_visibility', '—')}", muted_style))
    story.append(Spacer(1, 6))

    story.append(Paragraph("Strategic Improvements", h2_style))
    if rec.get("priority_action"):
        story.append(Paragraph(f"<b>Priority Action:</b> {rec['priority_action']}", body_style))
        story.append(Spacer(1, 4))
    if rec.get("headline_suggestions"):
        story.append(Paragraph(f"<b>Optimized Headline suggestion:</b> {rec['headline_suggestions'][0]}", body_style))

    doc.build(story)
    return buf.getvalue()
